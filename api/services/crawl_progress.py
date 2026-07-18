# -*- coding: utf-8 -*-
"""Count crawled comment records from output files."""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

from ..schemas import CrawlerStartRequest
from tools.save_path_utils import comment_file_globs, count_csv_rows


def resolve_save_base(project_root: Path, save_data_path: str) -> Path:
    raw = (save_data_path or "./data/comments").strip()
    if raw.startswith("./"):
        return (project_root / raw[2:]).resolve()
    if raw.startswith("/"):
        return Path(raw).resolve()
    return (project_root / raw).resolve()


def _unique_paths(paths: list[Path]) -> list[Path]:
    seen: set[str] = set()
    unique: list[Path] = []
    for path in paths:
        key = str(path.resolve())
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def list_comment_files(
    project_root: Path,
    config: CrawlerStartRequest,
    date_str: Optional[str] = None,
) -> list[Path]:
    """List all comment output files for the current save layout."""
    save_option = config.save_option.value
    if save_option in {"db", "sqlite", "mongodb"}:
        return []

    platform = config.platform.value
    crawler_type = config.crawler_type.value
    base = resolve_save_base(project_root, config.save_data_path)
    date_str = date_str or datetime.now().strftime("%Y-%m-%d")

    if save_option == "excel":
        if config.split_by_video:
            return _unique_paths([p for p in base.glob("**/*.xlsx") if p.is_file()])
        for folder in (base / platform, base / "bilibili"):
            if folder.is_dir():
                return list(folder.glob("*.xlsx"))
        return []

    ext_map = {"csv": "csv", "jsonl": "jsonl", "json": "json"}
    ext = ext_map.get(save_option)
    if not ext:
        return []

    if config.split_by_video:
        candidates: list[Path] = []
        for pattern in comment_file_globs(date_str, ext):
            candidates.extend(base.glob(pattern))
        return _unique_paths([p for p in candidates if p.is_file()])

    filename = f"{crawler_type}_comments_{date_str}.{ext}"
    direct = base / platform / ext / filename
    return [direct] if direct.is_file() else []


def snapshot_comment_files(
    project_root: Path,
    config: CrawlerStartRequest,
) -> dict[str, int]:
    """Record existing comment file sizes at session start."""
    snapshot: dict[str, int] = {}
    for path in list_comment_files(project_root, config):
        snapshot[str(path.resolve())] = count_records_in_file(path, config.save_option.value)
    return snapshot


def count_session_comment_progress(
    project_root: Path,
    config: CrawlerStartRequest,
    started_at: Optional[datetime],
    snapshot: Optional[dict[str, int]] = None,
) -> int:
    """Count comments written during the current session only."""
    snapshot = snapshot or {}
    session_start = started_at.replace(microsecond=0) if started_at else None
    total = 0

    for path in list_comment_files(project_root, config):
        resolved = str(path.resolve())
        file_mtime = datetime.fromtimestamp(path.stat().st_mtime)
        if session_start and file_mtime < session_start:
            continue

        current = count_records_in_file(path, config.save_option.value)
        baseline = snapshot.get(resolved, 0)
        if resolved not in snapshot and session_start and file_mtime >= session_start:
            baseline = 0
        elif current < baseline:
            # Output file was cleared and rewritten during this session.
            baseline = 0
        total += max(0, current - baseline)

    return total


def count_final_session_comments(
    project_root: Path,
    config: CrawlerStartRequest,
    started_at: Optional[datetime],
    snapshot: Optional[dict[str, int]] = None,
) -> int:
    """Count comments for the active session, preferring the latest output file."""
    session_count = count_session_comment_progress(
        project_root, config, started_at=started_at, snapshot=snapshot
    )
    active_file = resolve_comments_file(project_root, config, started_at=started_at)
    if not active_file:
        return session_count
    file_total = count_records_in_file(active_file, config.save_option.value)
    return max(session_count, file_total)


def _filter_session_candidates(
    candidates: list[Path],
    started_at: Optional[datetime],
) -> list[Path]:
    if not started_at or not candidates:
        return []
    session_start = started_at.replace(microsecond=0)
    return [
        path
        for path in candidates
        if datetime.fromtimestamp(path.stat().st_mtime) >= session_start
    ]


def resolve_comments_file(
    project_root: Path,
    config: CrawlerStartRequest,
    started_at: Optional[datetime] = None,
    snapshot: Optional[dict[str, int]] = None,
) -> Optional[Path]:
    """Resolve the primary comment file touched during this session."""
    save_option = config.save_option.value
    if save_option in {"db", "sqlite", "mongodb"}:
        return None

    base = resolve_save_base(project_root, config.save_data_path)

    if save_option == "excel":
        if config.split_by_video:
            candidates = list(base.glob("**/*.xlsx"))
        else:
            platform = config.platform.value
            candidates = []
            for folder in (base / platform, base / "bilibili"):
                if folder.is_dir():
                    candidates.extend(folder.glob("*.xlsx"))
        candidates = _filter_session_candidates(candidates, started_at)
        if not candidates:
            return None
        return max(candidates, key=lambda path: path.stat().st_mtime)

    candidates = list_comment_files(project_root, config)
    if config.split_by_video:
        candidates = _filter_session_candidates(candidates, started_at)
        if candidates:
            return max(candidates, key=lambda path: path.stat().st_mtime)
        return None

    if candidates:
        direct = candidates[0]
        if started_at and datetime.fromtimestamp(direct.stat().st_mtime) < started_at.replace(microsecond=0):
            return None
        return direct
    return None


def count_records_in_file(file_path: Path, save_option: str) -> int:
    if not file_path.is_file():
        return 0

    try:
        if save_option == "csv":
            counted = count_csv_rows(file_path, quick=False)
            return counted if counted is not None else 0

        if save_option == "jsonl":
            with open(file_path, "r", encoding="utf-8") as handle:
                return sum(1 for line in handle if line.strip())

        if save_option == "json":
            with open(file_path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if isinstance(data, list):
                return len(data)
            return 1 if data else 0

        if save_option == "excel":
            import openpyxl

            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            try:
                if "Comments" not in workbook.sheetnames:
                    return 0
                sheet = workbook["Comments"]
                max_row = sheet.max_row or 0
                return max(0, max_row - 1)
            finally:
                workbook.close()
    except Exception:
        return 0

    return 0
